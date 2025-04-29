import io
from typing import Dict, List
from openpyxl import load_workbook

from .base import FileProcessor


class XLSXProcessor(FileProcessor):
    @staticmethod
    def process_chunk(rows: List[List], columns: List[str]) -> Dict:
        rows = [dict(zip(columns, row)) for row in rows]
        return {
            "columns": columns,
            "rows": rows
        }

    @staticmethod
    def get_chunk_data(ws, start_row: int, chunk_size: int) -> List[List]:
        chunk = []
        for idx, row in enumerate(ws.iter_rows(min_row=start_row, max_row=start_row + chunk_size - 1)):
            chunk.append([str(cell.value) if cell.value is not None else "" for cell in row])
        return chunk

    async def process(self, content: bytes, filename: str) -> Dict:
        try:
            buffer = io.BytesIO(content)
            chunk_size = 1000  # Reduced chunk size
            pages = []
            page_counter = 1

            wb = load_workbook(buffer)

            for sheet in wb.sheetnames:
                ws = wb[sheet]
                # Get columns from first row
                columns = [str(cell.value) for cell in next(ws.rows)]

                # Get total rows
                total_rows = ws.max_row - 1  # Excluding header

                # Process chunks
                for start_row in range(2, total_rows + 2, chunk_size):
                    # Get chunk data directly
                    chunk_data = self.get_chunk_data(ws, start_row, chunk_size)
                    if not chunk_data:
                        continue

                    # Process chunk synchronously
                    result = self.process_chunk(chunk_data, columns)

                    pages.append({
                        "page_number": page_counter,
                        "sheet_name": sheet,
                        "text": "",
                        "tables": [result],
                        "images": []
                    })
                    page_counter += 1

            wb.close()
            return {
                "filename": filename,
                "pages": pages,
                "page_count": len(pages)
            }
        except Exception as e:
            return {"filename": filename, "error": f"Failed to process XLSX: {str(e)}"}
